# -*- coding: utf-8 -*-
# @Time    : 2018/9/11 13:13
# @Author  : FrankZ
# @Email   : FrankZ981210@gmail.com
# @File    : reactor_proxy_server
# @Software: PyCharm

from twisted.protocols.basic import LineReceiver
from twisted.internet import defer, main, threads
from twisted.internet.protocol import ServerFactory, ReconnectingClientFactory


from rpc_helper import ServerProxy, RPCProxy
from sql_helper import MsSqlServer, SqlService
import psycopg2

import struct
import configparser
import openpyxl
import time
import logging
from logging.handlers import TimedRotatingFileHandler
import os


# 扫描枪管理客户端传输协议，基于LineReceiver
class ProxyProtocol(LineReceiver):
    peer_ip = ""

    def connectionMade(self):
        self.peer_ip = self.transport.getPeer().host
        logger.info("扫描枪管理客户端连接：%s" % (self.transport.getPeer(),))

    def lineReceived(self, line):
        self._on_data_received(self.peer_ip, line)

    def _on_data_received(self, peer_ip, data):
        self.factory.data_receive(peer_ip, data)

    # 连接丢失时的回调
    def connectionLost(self, reason):
        logger.error("扫描枪管理客户端连接丢失：%s" % (self.transport.getPeer(),))
        self.transport.abortConnection()
        pass


# 当系统前时间
def now_time():
    return time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())


# 代理服务工厂，管理扫描枪客户端连接 与 结果服务反馈
class ProxyServerFactory(ServerFactory):
    protocol = ProxyProtocol
    result_protocol = None
    deferred = []

    output_panel = {}
    config = None

    def __init__(self):
        self.deferred = [self._init_set_deferred(), self._init_lost_deferred(), self._init_result_info_deferred()]
        self.config = config
        self._init_status_history()

    def startFactory(self):
        result_factory = ResultClientFactory(self.deferred)
        result_host = self.config["resultServer"]["result_host"]
        result_port = int(self.config["resultServer"]["result_port"])
        from twisted.internet import reactor
        reactor.connectTCP(result_host, result_port, result_factory)

    def stopFactory(self):
        logger.info("代理服务程序关闭")

    # 从扫描枪控制客户端接收数据，
    # peer_ip ：扫描枪控制客户端的ip
    def data_receive(self, peer_ip, data):
        try:
            _type, _source, _keys = struct.unpack("3s60s20s", data)
            input_type = _type.decode("utf-8").rstrip('\0')
            input_source = _source.decode("utf-8").rstrip('\0')
            input_keys = _keys.decode("utf-8").rstrip('\0')
            if '-' in input_keys:
                pre_key = input_keys.split('-')[0]
                key_data = input_keys.split('-')[1]
            else:
                pre_key = "000"
                key_data = input_keys
            if self.config.getboolean('data_prefix', 'isdigit'):
                if not pre_key.isdigit():
                    raise ValueError('前缀配置为数字，当前前缀：%s, 未处理，丢弃数据 : %s' % (pre_key, input_keys))
            pre_length = self.config.getint('data_prefix', 'prefix_length')
            if len(pre_key) > pre_length:
                raise ValueError('前缀配置长度为%s，当前前缀：%s 超出范围, 未处理，丢弃数据 : %s' % (pre_length, pre_key, input_keys))
            elif len(pre_key) < pre_length:
                pre_key = pre_key.zfill(pre_length)

            input_path = peer_ip+"&"+input_source+"&"+pre_key

            if self.config.has_option("Route", input_path):
                # 更新已被使用的配置
                # 已经为该路径配置路由，从预选列表中删除
                if self.config.has_option("input_path", input_path):
                    self.config.remove_option("input_path", input_path)
                    with open(real_path + "/config.conf", 'w', encoding="utf-8") as configfile:
                        self.config.write(configfile)
                if self.check_for_deal(input_path, input_keys):
                    logger.info(" %s-%s-已提交处理" % (input_path, input_keys))
                    self.deal_data(input_path,key_data,input_type)
                else:
                    logger.warning(" %s-%s-未处理" % (input_path, input_keys))

            elif self.config.has_option("input_path", input_path):
                logger.warning(" %s - 未设置结果展示的panel " % input_path)
            else:
                logger.info(" %s - 新输入路径 : %s" % (input_path, input_keys))
                self.config["input_path"][input_path] = "#"
                with open(real_path + "/config.conf", 'w', encoding="utf-8") as configfile:
                    self.config.write(configfile)
        except Exception as e:
            logger.error(" 数据接收时发生异常 - data_receive : %s" % e.__repr__())

    def check_for_deal(self, _path, _keys):
        if self.config.getboolean("proxyServer", "force_deal"):
            logger.warning(" %s-收到条码- %s - 强制处理 - force_deal:True" % (_path, _keys))
            return True
        else:
            output_s = self.config["Route"][_path].split('/')
            for o in output_s:
                if o in self.output_panel:
                    logger.info(" %s-收到条码-%s-禁止强制-force_deal:False- 发现可用的panel" % (_path, _keys))
                    return True
            logger.warning(" %s-收到条码-%s-禁止强制-force_deal:False- 无可用panel" % (_path, _keys))
            return False

    def send_msg_to(self, tar):
        o_s = self.config["Route"][tar[0]].split('/')
        for o in o_s:
            if o in self.output_panel:
                # 记录该输入路径
                if tar[0] not in self.output_panel[o]:
                    self.output_panel[o].append(tar[0])
                # 向配置的面板发送消息
                self.send_result(o+"&"+tar[1])
                logger.info(" 已发送消息 - panel: %s-%s" % (o, tar[1]))
            else:
                logger.warning(" 未使用 - panel: %s " % o)

    login_recode = {}
    state_recode = {}
    num_recode = {}
    singal_num_recode = {}
    status_history = {}
    scan_history = {}
    picking_queue = {}
    picking_defer = {}

    # 登录rpc检测
    def login(self, _path, username, password):
        rpc_host = self.config["Odoo"]["rpc_host"]
        rpc_port = self.config["Odoo"]["rpc_port"]
        rpc_db = self.config["Odoo"]["rpc_db"]
        odoologinurl = ("http://%s:%s/xmlrpc/2/common" % (rpc_host, rpc_port))
        common = ServerProxy(odoologinurl)
        try:
            uid = common.authenticate(rpc_db, username, password, {})
        except ConnectionRefusedError as cre:
            logger.info("-%s-登录失败-rpc连接失败-%s-%s-%s" % (_path, rpc_host, rpc_port, cre.__repr__()))
            return False
        if uid:
            rpc = RPCProxy(uid, password, host=rpc_host, port=rpc_port, dbname=rpc_db)
            return_msg = rpc('res.users', 'search_read', [('id', '=', uid)], ['name'])

            self.login_recode[_path] = (return_msg[0].get('name'), username, uid, password)
            self.num_recode[_path] = 0
            self.singal_num_recode[_path] = 0
            self.state_recode[_path] = None
            self.scan_history[_path] = []

            logger.info("-%s-登录成功-用户-%s:%s" % (_path, self.login_recode[_path][0], self.login_recode[_path][1]))
            return True
        else:
            logger.info("-%s-登录失败-登录条码错误-%s" % (_path, username))
            return False

    # 登出
    def login_out(self, _path):
        login_value = self.login_recode.get(_path, None)
        if login_value:
            logger.info("-%s-退出成功-用户-%s:%s" % (_path, self.login_recode[_path][0], self.login_recode[_path][1]))
            self.login_recode[_path] = None
            self.state_recode[_path] = None
            self.num_recode[_path] = 0
            self.singal_num_recode[_path] = 0
            self.scan_history[_path] = []
            return True
        else:
            return False

    # 判断是否为登录条码
    def is_login_value(self, _path, value):
        if value and value.startswith("9999"):
            if (_path in self.picking_queue) and self.picking_queue[_path]:
                logger.info("-%s-禁止切换登录，%s 正在处理中" % (_path, self.state_recode[_path][0]))
                return True
            else:
                return self.login(_path, value[-6:], value[-6:][::-1])
        else:
            return False

    # 判断是否为登出条码
    def is_login_out_value(self, _path, value):
        if value and value.startswith("9998"):
            return self.login_out(_path)
        else:
            return False

    # 初始化状态条码
    def _init_status_history(self):
        try:
            wb = openpyxl.load_workbook(real_path + "/status.xlsx", read_only=True)
            sheet = wb[wb.sheetnames[0]]
            rows = sheet.iter_rows(min_row=2, max_col=2)
            for row in rows:
                self.status_history[row[0].value.replace(' ', '').upper()] = row[1].value.strip(' ').replace('&', '').replace(':', '')
            logger.info("初始化状态列表")
            wb.close()
            return True
        except Exception as e:
            logger.info("初始化状态列表失败-%s" % e.__repr__())
            return False
        pass

    def check_status(self, _path, status):
        status = status.upper()
        if status in self.status_history:
            self.state_recode[_path] = (self.status_history[status], status)
            self.num_recode[_path] = 0
            self.singal_num_recode[_path] = 0
            logger.info("-%s-设置状态成功-%s:%s" % (_path, self.state_recode[_path][0], status))
            return True
        return False

    # 判断是否为状态条码
    def is_state_value(self, _path, value):
        if len(value) < 8:
            return self.check_status(_path, value)
        else:
            return False

    # 判断是否为生产单号
    def is_job_num(self, value):
        if len(value) == 10:
            return True
        else:
            return False

    # 判断是否为出库设置条码
    def is_picking_function(self, _path, fnc):
        if fnc == '03' or fnc == '04':
            if _path not in self.picking_queue:
                self.picking_queue[_path] = []
                self.picking_defer[_path] = None
            if self.picking_queue[_path]:
                logger.info("-%s-禁止切换出库条码，%s 正在处理中" % (_path, self.state_recode[_path][0]))
                return True
            if fnc == '03':
                self.state_recode[_path] = ('[O片O] 出库', 'lens')
                self.singal_num_recode[_path] = 0
                logger.info("-%s-设置镜片出库成功-%s:%s" % (_path, self.state_recode[_path][0], fnc))
                return True
            elif fnc == '04':
                self.state_recode[_path] = ('[口架口] 出库', 'frame')
                self.singal_num_recode[_path] = 0
                logger.info("-%s-设置镜架出库成功-%s:%s" % (_path, self.state_recode[_path][0], fnc))
                return True
        else:
            return False

    # 判断是否清零条码
    def is_clear_function(self, _path, fnc):
        if fnc == '02':
            self.singal_num_recode[_path] = 0
            logger.info("-%s-执行数量清零 - %s" % (_path, fnc))
            return True
        return False

    # 判断是否外协收料条码
    def is_out_source_pick_function(self, _path, fnc):
        if fnc == '05':
            self.state_recode[_path] = ('外协收料', '05')
            self.num_recode[_path] = 0
            self.singal_num_recode[_path] = 0
            logger.info("-%s-设置外协收料成功-%s:%s" % (_path, self.state_recode[_path][0], fnc))
            return True
        return False

    # 返回值前缀说明：
    #     login:登录信息,数据格式为："login:用户名:工号"
    #     num:扫描总数,数据格式为："num:数量"
    #     state:状态名,数据格式为："state:状态名"
    #     cls:清屏,数据格式为："cls:1"
    #     info:通知信息,数据格式为："info:通知信息"
    #     error: 错误信息,数据格式为："error:错误信息"
    #     msg:一般返回消息,数据格式为："msg:一般消息"
    def deal_data(self, _path, _data, data_type):
        login_value = self.login_recode.get(_path, None)
        if len(_data) <= 0:
            msg = "error:扫描有误"
        elif self.is_login_out_value(_path, _data):
            msg = "login:%s:%s&num:0&state:未扫状态条码&cls:1&info:退出成功" % ("未登录", " ", )
        elif self.is_login_value(_path, _data):
            if (_path in self.picking_queue) and self.picking_queue[_path]:
                msg = "error:禁止切换登录，还存在未处理出库单，请等待完成！"
            else:
                msg = "login:%s:%s&num:0&state:未扫状态条码&cls:1&info:登录成功" % \
                   (self.login_recode[_path][0], self.login_recode[_path][1])
        elif self.is_state_value(_path, _data):
            # 状态录入设置
            if login_value:
                msg = "state:%s&num:0&cls:1&info:设置%s成功" % (self.state_recode[_path][0], self.state_recode[_path][0])
            else:
                msg = "error:未登录"
        elif self.is_out_source_pick_function(_path, _data):
            # 外协收料设置
            if login_value:
                msg = "state:%s&num:0&cls:1&info:设置%s成功" % (self.state_recode[_path][0], self.state_recode[_path][0])
            else:
                msg = "error:未登录"
        elif self.is_clear_function(_path, _data):
            # 清除设置
            if data_type == 'pic':
                msg = "state:数量:0-%s&cls:1&info:数量清除成功" % self.state_recode[_path][0]
            else:
                msg = "state:%s&cls:1&info:数量清除成功" % self.state_recode[_path][0]
        elif data_type == 'pic' and self.is_picking_function(_path, _data):
            # 出库设置
            if login_value:
                if self.picking_queue[_path]:
                    msg = "error:禁止切换出库条码，还存在未处理出库单，请等待完成！"
                else:
                    msg = "state:数量:0-%s&num:%s&cls:1&info:设置%s成功" % (self.state_recode[_path][0], self.num_recode[_path], self.state_recode[_path][0])
            else:
                msg = "error:未登录"
        elif self.is_job_num(_data):
            # 根据设置处理生产单号
            state_value = self.state_recode.get(_path, None)
            if login_value is None:
                msg = "error:未登录"
            elif state_value is None:
                msg = "error:未扫状态条码"
            else:
                msg = ""
                if state_value[1] == '05':
                    d = threads.deferToThread(self.out_source_pick_scan, _path, _data)
                    d.addCallback(self.send_msg_to)
                elif data_type == 'pic' and (state_value[1] in ['frame', 'lens']):
                    self.picking_queue[_path].append(_data)
                    logger.info(" %s - 出库操作 - 加入延迟队列 - %s" % (_path, _data))
                    if not self.picking_defer[_path]:
                        logger.info(" %s - 出库操作 - 启动出库执行" % _path)
                        from twisted.internet import reactor
                        reactor.callLater(0.01, self.pickingTask, _path)
                else:
                    d = threads.deferToThread(self.pickingbusinessforstatus, _path, _data)
                    d.addCallback(self.send_msg_to)
        else:
            # 其他条码
            msg = "error:"+now_time() + "-" + _data + "-" + "条码未设置使用"
        if msg:
            self.send_msg_to((_path, msg))

    def out_source_pick_scan(self, _path, jobnum):
        """
        外协收料逻辑
        :param _path:
        :param jobnum:
        :return:
        """
        success = False
        login_value = self.login_recode[_path]
        state_value = self.state_recode[_path]
        try:
            host = self.config.get("Odoo", "rpc_host")
            port = self.config.get("Odoo", "rpc_port")
            db = self.config.get("Odoo", "rpc_db")

            uid = self.login_recode[_path][2]
            pwd = self.login_recode[_path][3]
            rpc = RPCProxy(uid, pwd, host=host, port=port, dbname=db)
            logger.error(" %s-%s:执行未来收料批次" % (_path, jobnum))
            # 获取生产单(制造单)
            ids = rpc("mrp.production", "search", [("state", "not in", ["cancel", "done"]),
                                                   ["restrict_lot_id.name", "=", jobnum]])
            # 获取采购单
            if ids:
                PurchaseOrderRecords = rpc("purchase.order", "search_read", [("mo_id.id", "=", ids[0])], ["name"])
                if PurchaseOrderRecords:
                    strSucceed = rpc("stock.picking", "action_done_remote2", jobnum, 'lens picking for production')
                    if strSucceed == "success!":
                        msg = " 来料接收:操作成功"
                        success = True
                        self.num_recode[_path] += 1
                        self.singal_num_recode[_path] += 1
                        logger.info(" %s-%s:%s" % (_path, jobnum, msg))
                    else:
                        msg = jobnum + " 来料接收:操作失败 : " + strSucceed
                        logger.error(" %s-%s:%s" % (_path, jobnum, msg))
                else:
                    msg = " :无外协信息,请联系主管,确认采购询价单"
                    logger.error(" %s-%s:获取采购单失败" % (_path, jobnum))
            else:
                msg = " :无外协信息,请联系主管,确认制造单"
                logger.error(" %s-%s:获取生产单(制造单)失败" % (_path, jobnum))
        except:
            msg = " :连接服务器失败，请重试！"
            success = False
            logger.error(" %s-%s:连接服务器失败" % (_path, jobnum))
        view_msg = now_time() + "-" + jobnum + "-" + self.state_recode[_path][0] + "-" + msg
        if success:
            return _path, "login:%s:%s&state:%s&num:%d&msg:%s" % \
                   (login_value[0], login_value[1], state_value[0], self.num_recode[_path], view_msg)
        else:
            return _path, "login:%s:%s&state:%s&num:%d&error:%s" % \
                   (login_value[0], login_value[1], state_value[0], self.num_recode[_path], view_msg)

    def pickingTask(self, _path):
        """
        出库任务队列
        :return:
        """
        if _path in self.picking_queue:
            if len(self.picking_queue[_path]) > 0:
                _data = self.picking_queue[_path].pop(0)
                logger.info(" %s - 出库操作 - 执行出库 - %s" % (_path, _data))
                self.picking_defer[_path] = threads.deferToThread(self.pickingbusiness, _path, _data)
                self.picking_defer[_path].addCallback(self.pickingTask)
            else:
                logger.info(" %s - 出库操作 - 队列已空 - 关闭执行" % _path)
                self.picking_defer[_path] = None

    def pickingbusiness(self, _path, jobnum):
        """
        出库逻辑
        :param _path:
        :param jobnum:
        :return:
        """
        try:
            host = self.config.get("Odoo", "rpc_host")
            port = self.config.get("Odoo", "rpc_port")
            db = self.config.get("Odoo", "rpc_db")

            uid = self.login_recode[_path][2]
            password = self.login_recode[_path][3]
            rpc = RPCProxy(uid, password, host=host, port=port, dbname=db)
            returnmsg = rpc('mrp.production', 'rpc_action_picking_done', jobnum, self.state_recode[_path][1])
            # 记录出库历史
            if returnmsg == '100':
                if self.config.getboolean("Odoo", "record_pick_history"):
                    try:
                        db_host = self.config.get("OdooPgSql", "host")
                        db_port = int(self.config.get("OdooPgSql", "port"))
                        db_usr = self.config.get("OdooPgSql", "user")
                        db_pwd = self.config.get("OdooPgSql", "pwd")
                        db_db = self.config.get("OdooPgSql", "db")
                        history_db = SqlService(util=psycopg2,host=db_host,port=db_port,user=db_usr,pwd=db_pwd,db=db_db)
                        sql = "insert into picking_history(job_num,uid,picking_time,picking_type) values ('%s','%s','%s','%s')"\
                              % (jobnum,uid,now_time(),self.state_recode[_path][1])
                        history_db.ExecNonQuery(sql)
                        history_db.CloseDB()
                        logger.warning('保存出库历史成功 - %s' % jobnum)
                    except Exception as e:
                        logger.warning('保存出库历史失败 - %s : %s' % (jobnum, e.__repr__()))
            self.send_msg_to(self.build_picking_msg(_path, jobnum, returnmsg))
        except Exception as e:
            self.send_msg_to(self.build_picking_msg(_path, jobnum, "777"))
        return _path

    # 上传扫描订单的状态，_path为扫描来源，jobnum为扫描的生产单号
    def pickingbusinessforstatus(self, _path, jobnum):
        """
        状态扫描逻辑
        :param _path:
        :param jobnum:
        :return:
        """
        try:
            host = self.config.get("Odoo", "rpc_host")
            port = self.config.get("Odoo", "rpc_port")
            db = self.config.get("Odoo", "rpc_db")
            mssqlhost = self.config["MsSql"]["mssqlhost"]
            mssqldb = self.config["MsSql"]["mssqldb"]
            mssqluser = self.config["MsSql"]["mssqluser"]
            mssqlpwd = self.config["MsSql"]["mssqlpwd"]
            statusreminder = self.config["Odoo"]["statusreminder"]
            maxHistory = self.config["Odoo"]["maxHistory"]

            uid = self.login_recode[_path][2]
            password = self.login_recode[_path][3]

            if jobnum in self.scan_history[_path]:
                # 重复扫描
                strrepeatscanning = "999"
                return self.build_status_msg(_path, jobnum, strrepeatscanning)
            else:
                # RPC 上传状态
                statuscode = self.state_recode[_path][1]
                rpc = RPCProxy(uid, password, host=host, port=port, dbname=db)
                info = []
                info.append(jobnum)
                info.append(statuscode)
                info.append(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())))

                # statuscode="REHC"
                if statusreminder.find(statuscode) == -1:
                    info.append("0")
                else:
                    info.append("1")

                returnmsg = rpc('mrp.work.flow', 'flow_create', info)
                # 更新旧数据库中数据
                objectmssql = MsSqlServer(host=mssqlhost, user=mssqluser, pwd=mssqlpwd, db=mssqldb)
                sqlinsertstatus = """ UPDATE [dbo].[order] set [update_date] = '%s',[status_code] = '%s' WHERE [job_num]='%s'""" % (
                    time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())), statuscode, jobnum
                )
                sql = sqlinsertstatus
                objectmssql.ExecNonQuery(sql)
                # returnmsg = "100"
                # 成功的状态被标记，不能重复上传
                if returnmsg == "100":
                    history_list = self.scan_history[_path]
                    if len(history_list) == maxHistory:
                        history_list.pop(0)
                    history_list.append(jobnum)
                return self.build_status_msg(_path, jobnum, returnmsg)
        except Exception as e:
            # print(e.__repr__())
            return self.build_status_msg(_path, jobnum, "777")

    def pickingbusinessforprint(self, _path, _data):
        pass

    # 构建显示消息，按照一定的格式构建，消息格式参照 deal_data 消息前缀说明
    def build_status_msg(self, _path, jobnum, returnmsg):
        login_value = self.login_recode[_path]
        state_value = self.state_recode[_path]
        if returnmsg == 'SHIP':
            returnmsg = "状态同步失败-订单已完成或取消"
            msg = now_time() + "-" + jobnum + "-" + self.state_recode[_path][0] + "-" + returnmsg
            return _path, "login:%s:%s&state:%s&num:%d&error:%s" % \
                   (login_value[0], login_value[1], state_value[0], self.num_recode[_path], msg)
        elif returnmsg != '100':

            if returnmsg == '502':
                returnmsg = "状态同步失败-工单不存在"
            elif returnmsg == '505':
                returnmsg = "状态同步失败-订单不存在"
            elif returnmsg == '600':
                returnmsg = self.state_recode[_path][0] + "已返工三次及以上"
            elif returnmsg == '999':
                returnmsg = "状态已同步-重复扫描"
            else:
                returnmsg = "状态同步失败-异常"

            msg = now_time() + "-" + jobnum + "-" + self.state_recode[_path][0] + "-" + returnmsg
            return _path, "login:%s:%s&state:%s&num:%d&error:%s" % \
                   (login_value[0], login_value[1], state_value[0], self.num_recode[_path], msg)

        else:
            returnmsg = "状态同步成功"
            self.num_recode[_path] += 1
            self.singal_num_recode[_path] += 1
            msg = now_time() + "-" + jobnum + "-" + self.state_recode[_path][0] + "-" + returnmsg

            return _path, "login:%s:%s&state:%s&num:%d&msg:%s" % \
                   (login_value[0], login_value[1], state_value[0], self.num_recode[_path], msg)

    def build_picking_msg(self, _path, jobnum, returnmsg):
        login_value = self.login_recode[_path]
        state_value = self.state_recode[_path]
        if returnmsg != '100':

            if returnmsg == "501":
                returnmsg = "镜片已出库"
            if returnmsg == "502":
                returnmsg = "镜架已出库"
            elif returnmsg == "503":
                returnmsg = "无库存"
            elif returnmsg == "504":
                returnmsg = "保留异常"
            elif returnmsg == "505":
                returnmsg = "出库动作异常"
            elif returnmsg == "506":
                returnmsg = "出库单不存在"
            elif returnmsg == "507":
                returnmsg = "生产单不存在"
            elif returnmsg == "508":
                returnmsg = "不需要出库"
            else:
                returnmsg = "出库失败-异常"

            msg = now_time() + "-" + jobnum + "-" + self.state_recode[_path][0] + "-" + returnmsg
            return _path, "login:%s:%s&state:数量:%s-%s&num:%d&error:%s" % \
                   (login_value[0], login_value[1], self.singal_num_recode[_path], state_value[0], self.num_recode[_path], msg)

        else:
            returnmsg = "出库成功"
            self.num_recode[_path] += 1
            self.singal_num_recode[_path] += 1
            msg = now_time() + "-" + jobnum + "-" + self.state_recode[_path][0] + "-" + returnmsg

            return _path, "login:%s:%s&state:数量:%s-%s&num:%d&msg:%s" % \
                   (login_value[0], login_value[1], self.singal_num_recode[_path], state_value[0], self.num_recode[_path], msg)


    # 接收结果显示服务器消息的回调，结果显示服务器向代理服务注册其控制的展示面板
    # 注册格式为："r&面板类型&面板唯一路径"
    # 取消注册格式为："ur&面板类型&面板唯一路径"
    # 面板唯一路径："面板ip;面板名"
    def on_result_info_receive(self, data):
        self.deferred[2] = self._init_result_info_deferred()
        try:
            _op, _panel = data.split('&', maxsplit=1)
            if _op == 'r':
                self.output_panel[_panel] = []
                # 更新配置文件
                # if not self.config.has_option("outer_panel",_panel):
                #     self.config["outer_panel"][_panel] = "#"
                #     with open(real_path + "/config.conf", 'w', encoding="utf-8") as configfile:
                #         self.config.write(configfile)
                logger.info("panel注册 : %s" % _panel)
            elif _op == 'ur':
                for _path in self.output_panel[_panel]:
                    self.login_recode[_path] = None
                    self.state_recode[_path] = None
                    self.num_recode[_path] = 0
                    self.singal_num_recode[_path] = 0
                    self.scan_history[_path] = []
                del self.output_panel[_panel]
                logger.info("panel 注销 : %s" % _panel)
            else:
                raise Exception("错误指令 : %s " % _op)
        except Exception as e:
            logger.error("处理panel时发生错误: %s " % e.__repr__())

    def _init_set_deferred(self):
        d = defer.Deferred()
        d.addCallbacks(self.set_result_protocol, self.set_result_failed)
        return d

    def _init_lost_deferred(self):
        d = defer.Deferred()
        d.addBoth(self.on_result_protocol_lost)
        return d

    def _init_result_info_deferred(self):
        d = defer.Deferred()
        d.addBoth(self.on_result_info_receive)
        return d

    def set_result_protocol(self, p):
        self.deferred[0] = self._init_set_deferred()
        self.result_protocol = p

    def set_result_failed(self, err):
        self.deferred[0] = self._init_set_deferred()

    def on_result_protocol_lost(self, err):
        self.deferred[1] = self._init_lost_deferred()
        self.output_panel = {}
        self.result_protocol = None

    # 向结果服务器发送结果的统一接口
    # 结果格式为："面板类型&面板唯一路径&消息"
    # 消息格式参照 deal_data 消息前缀说明
    def send_result(self, result):
        if self.result_protocol is None:
            logger.error("Result server 未连接")
        else:
            from twisted.internet import reactor
            reactor.callLater(0.01, self.result_protocol.send_result, result)


# 与ResultServer通信Protocol
class ResultClientProtocol(LineReceiver):

    def connectionMade(self):
        self.factory.connect_success(self)

    def lineReceived(self, line):
        self.factory.data_receive(line.decode())

    def send_result(self, result):
        self.sendLine(result.encode())

    def connectionLost(self, reason=main.CONNECTION_LOST):
        self.transport.loseConnection()


# 与ResultServer通信Factory
class ResultClientFactory(ReconnectingClientFactory):
    p = None
    maxDelay = 1800

    def __init__(self, deferred):
        self.deferred = deferred

    def buildProtocol(self, addr):
        self.p = ResultClientProtocol()
        self.p.factory = self
        self.resetDelay()
        return self.p

    def data_receive(self, data):
        self.deferred[2].callback(data)

    def connect_success(self, p):
        self.deferred[0].callback(p)
        logger.info("连接Result Server成功")

    def clientConnectionFailed(self, connector, reason):
        logger.error("连接Result Server失败, 等待重连...")
        ReconnectingClientFactory.clientConnectionFailed(self, connector, reason)

    def clientConnectionLost(self, connector, unused_reason):
        ReconnectingClientFactory.clientConnectionLost(self, connector, unused_reason)
        logger.error("丢失Result Server连接, 等待重连...")
        self.deferred[1].errback(unused_reason)


def main():
    logger.info("代理服务（Proxy Server）启动")
    factory = ProxyServerFactory()
    from twisted.internet import reactor
    proxy_port = int(config["proxyServer"]["proxy_port"])
    port = reactor.listenTCP(proxy_port, factory)
    logger.info('Proxy Serving 监听端口（proxy_port）： %d' % port.getHost().port)
    from twisted.internet import reactor
    reactor.suggestThreadPoolSize(25)
    reactor.run()


logger = logging.getLogger('Proxy')
logger.setLevel(logging.DEBUG)

real_path = os.path.split(os.path.realpath(__file__))[0]
fn = real_path + '/log/log.log'

fh = TimedRotatingFileHandler(fn, when='D', interval=1, backupCount=10, encoding='utf-8')
fh.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
fh.setFormatter(formatter)
logger.addHandler(fh)

config = configparser.ConfigParser(delimiters='=')
config.read(real_path + "/config.conf", encoding="utf-8")

if __name__ == '__main__':
    main()
