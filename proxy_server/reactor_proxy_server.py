# -*- coding: utf-8 -*-
# @Time    : 2018/9/11 13:13
# @Author  : FrankZ
# @Email   : FrankZ981210@gmail.com
# @File    : reactor_proxy_server
# @Software: PyCharm

from twisted.protocols.basic import LineReceiver
from twisted.internet import defer, main
from twisted.internet.protocol import ServerFactory, ReconnectingClientFactory


from rpc_helper import ServerProxy, RPCProxy
from sql_helper import MsSqlServer

import struct
import configparser
import openpyxl
import time
import logging
import os


# 扫描枪管理客户端传输协议，基于LineReceiver
class ProxyProtocol(LineReceiver):
    peer_ip = ""

    def connectionMade(self):
        self.peer_ip = self.transport.getPeer().host
        logger.info("扫描枪管理客户端连接：%s" % (self.transport.getPeer(),))

    def lineReceived(self, line):
        self._on_data_received(self.peer_ip, line)

    def _on_data_received(self, peer_ip, data):
        self.factory.data_receive(peer_ip, data)

    # 连接丢失时的回调
    def connectionLost(self, reason):
        logger.error("扫描枪管理客户端连接丢失：%s" % (self.transport.getPeer(),))
        self.transport.loseConnection()
        pass


# 当系统前时间
def now_time():
    return time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())


# 代理服务工厂，管理扫描枪客户端连接 与 结果服务反馈
class ProxyServerFactory(ServerFactory):
    protocol = ProxyProtocol
    result_protocol = None
    deferred = []

    output_panel = {}
    config = None

    def __init__(self):
        self.deferred = [self._init_set_deferred(), self._init_lost_deferred(), self._init_result_info_deferred()]
        self.config = configparser.ConfigParser(delimiters='=')
        self.config.read(os.path.split(os.path.realpath(__file__))[0] + "/config.conf", encoding="utf-8")
        # print("create %s" % self.deferred)
        self._init_status_history()

    def startFactory(self):
        result_factory = ResultClientFactory(self.deferred)
        from twisted.internet import reactor
        reactor.connectTCP("127.0.0.1", 6690, result_factory)

    def stopFactory(self):
        logger.info("代理服务程序关闭")

    # 从扫描枪控制客户端接收数据，
    # peer_ip ：扫描枪控制客户端的ip
    def data_receive(self, peer_ip, data):
        try:
            _type, _source, _keys = struct.unpack("3s60s20s", data)
            input_type = _type.decode("utf-8").rstrip('\0')
            input_source = _source.decode("utf-8").rstrip('\0')
            input_keys = _keys.decode("utf-8").rstrip('\0')

            input_path = peer_ip+"&"+input_source

            if self.config.has_option(peer_ip, input_path):
                o = self.config[peer_ip][input_path]
                if o == '#':
                    logger.warning("-%s-It isn't set panel to show " % input_path)
                elif o in self.output_panel:
                    # 记录该输入路径
                    if input_path not in self.output_panel[o]:
                        self.output_panel[o].append(input_path)
                    # 向配置的面板发送消息
                    msg = self.deal_data(input_path, input_keys)
                    logger.info("-%s-收到条码-%s-%s" % (input_path, input_keys, msg))
                    self.send_msg_to(o, msg)
                else:
                    logger.warning("result panel-%s- is useless" % o)
            else:
                logger.info("-%s-new path input " % input_path)

                if self.config.has_section(peer_ip):
                    self.config[peer_ip][input_path] = "#"
                else:
                    self.config[peer_ip] = {}
                    self.config[peer_ip][input_path] = "#"

                with open('config.conf', 'w', encoding="utf-8") as configfile:
                    self.config.write(configfile)
        except Exception as e:
            print(e.__repr__())
            pass

    def send_msg_to(self, target, msg):
        self.send_result(target+"&"+msg)

    login_recode = {}
    state_recode = {}
    num_recode = {}
    status_history = {}
    scan_history = {}

    # 登录rpc检测
    def login(self, _path, username, password):
        rpc_host = self.config["Odoo"]["rpc_host"]
        rpc_port = self.config["Odoo"]["rpc_port"]
        rpc_db = self.config["Odoo"]["rpc_db"]
        odoologinurl = ("http://%s:%s/xmlrpc/2/common" % (rpc_host, rpc_port))
        common = ServerProxy(odoologinurl)
        try:
            uid = common.authenticate(rpc_db, username, password, {})
        except ConnectionRefusedError as cre:
            logger.info("-%s-登录失败-rpc连接失败-%s-%s-%s" % (_path, rpc_host, rpc_port, cre.__repr__()))
            return False
        if uid:
            rpc = RPCProxy(uid, password, host=rpc_host, port=rpc_port, dbname=rpc_db)
            return_msg = rpc('res.users', 'search_read', [('id', '=', uid)], ['name'])

            self.login_recode[_path] = (return_msg[0].get('name'), username, uid, password)
            self.num_recode[_path] = 0
            self.state_recode[_path] = None
            self.scan_history[_path] = []

            logger.info("-%s-登录成功-用户-%s:%s" % (_path, self.login_recode[_path][0], self.login_recode[_path][1]))
            return True
        else:
            logger.info("-%s-登录失败-登录条码错误-%s" % (_path, username))
            return False

    # 登出
    def login_out(self, _path):
        login_value = self.login_recode.get(_path, None)
        if login_value:
            logger.info("-%s-退出成功-用户-%s:%s" % (_path, self.login_recode[_path][0], self.login_recode[_path][1]))
            self.login_recode[_path] = None
            self.state_recode[_path] = None
            self.num_recode[_path] = 0
            self.scan_history[_path] = []
            return True
        else:
            return False

    # 判断是否为登录条码
    def is_login_value(self, _path, value):
        if value and value.startswith("9999"):
            return self.login(_path, value[-6:], value[-6:][::-1])
        else:
            return False

    # 判断是否为登出条码
    def is_login_out_value(self, _path, value):
        if value and value.startswith("9998"):
            return self.login_out(_path)
        else:
            return False

    # 初始化状态条码
    def _init_status_history(self):
        try:
            wb = openpyxl.load_workbook(os.path.split(os.path.realpath(__file__))[0] + "/status.xlsx", read_only=True)
            sheet = wb[wb.sheetnames[0]]
            rows = sheet.iter_rows(min_row=2, max_col=2)
            for row in rows:
                self.status_history[row[0].value] = row[1].value
            logger.info("初始化状态列表")
            return True
        except Exception as e:
            logger.info("初始化状态列表失败-%s" % e.__repr__())
            return False
        pass

    def check_status(self, _path, status):
        if status in self.status_history:
            self.state_recode[_path] = (self.status_history[status], status)
            self.num_recode[_path] = 0
            logger.info("-%s-设置状态成功-%s:%s" % (_path, self.state_recode[_path][0], status))
            return True
        try:
            wb = openpyxl.load_workbook(os.path.split(os.path.realpath(__file__))[0] + "/status.xlsx", read_only=True)
            sheet = wb[wb.sheetnames[0]]
            rows = sheet.iter_rows(min_row=len(self.status_history)+1, max_col=2)
            for row in rows:
                self.status_history[row[0].value] = row[1].value
                for c in row:
                    if status == c.value:
                        self.state_recode[_path] = (sheet.cell(row=c.row, column=2).value, c.value)
                        self.num_recode[_path] = 0
                        logger.info("-%s-设置状态成功-%s:%s" % (_path, self.state_recode[_path][0], status))
                        wb.close()
                        return True
            wb.close()
            return False
        except Exception as e:
            return False

    # 判断是否为状态条码
    def is_state_value(self, _path, value):
        if len(value) < 8:
            return self.check_status(_path, value)
        else:
            return False

    # 返回值前缀说明：
    #     login:登录信息,数据格式为："login:用户名:工号"
    #     num:扫描总数,数据格式为："num:数量"
    #     state:状态名,数据格式为："state:状态名"
    #     cls:清屏,数据格式为："cls:1"
    #     info:通知信息,数据格式为："info:通知信息"
    #     error: 错误信息,数据格式为："error:错误信息"
    #     msg:一般返回消息,数据格式为："msg:一般消息"
    def deal_data(self, _path, _data):
        login_value = self.login_recode.get(_path, None)
        if len(_data) <= 0:
            return "error:扫描有误"
        elif self.is_login_out_value(_path, _data):
            return "login:%s:%s&num:0&state:未扫状态条码&cls:1&info:退出成功" % ("未登录", " ", )
        elif self.is_login_value(_path, _data):
            return "login:%s:%s&num:0&state:未扫状态条码&cls:1&info:登录成功" % \
                   (self.login_recode[_path][0], self.login_recode[_path][1])
        elif self.is_state_value(_path, _data):
            if login_value:
                return "state:%s&num:0&cls:1&info:设置状态成功" % self.state_recode[_path][0]
            else:
                return "error:未登录"
        else:
            state_value = self.state_recode.get(_path, None)
            if login_value is None:
                return "error:未登录"
            elif state_value is None:
                return "error:未扫状态条码"
            else:
                if self.config["Odoo"]["titleType"].split(';')[0] == '1':
                    return self.pickingbusiness(_path, _data)
                elif self.config["Odoo"]["titleType"].split(';')[0] == '2':
                    return self.pickingbusinessforstatus(_path, _data)
                else:
                    return self.pickingbusinessforprint(_path, _data)

    def pickingbusiness(self, _path, _data):
        pass

    # 上传扫描订单的状态，_path为扫描来源，jobnum为扫描的生产单号
    def pickingbusinessforstatus(self, _path, jobnum):
        try:
            host = self.config.get("Odoo", "rpc_host")
            port = self.config.get("Odoo", "rpc_port")
            db = self.config.get("Odoo", "rpc_db")
            mssqlhost = self.config["MsSql"]["mssqlhost"]
            mssqldb = self.config["MsSql"]["mssqldb"]
            mssqluser = self.config["MsSql"]["mssqluser"]
            mssqlpwd = self.config["MsSql"]["mssqlpwd"]
            statusreminder = self.config["Odoo"]["statusreminder"]
            maxHistory = self.config["Odoo"]["maxHistory"]

            uid = self.login_recode[_path][2]
            password = self.login_recode[_path][3]

            if jobnum in self.scan_history[_path]:
                # 重复扫描
                strrepeatscanning = "999"
                return self.build_status_msg(_path, jobnum, strrepeatscanning)
            else:
                # RPC 上传状态
                statuscode = self.state_recode[_path][1]
                rpc = RPCProxy(uid, password, host=host, port=port, dbname=db)
                info = []
                info.append(jobnum)
                info.append(statuscode)
                info.append(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())))

                # statuscode="REHC"
                if statusreminder.find(statuscode) == -1:
                    info.append("0")
                else:
                    info.append("1")

                returnmsg = rpc('mrp.work.flow', 'flow_create', info)
                # 更新旧数据库中数据
                objectmssql = MsSqlServer(host=mssqlhost, user=mssqluser, pwd=mssqlpwd, db=mssqldb)
                sqlinsertstatus = """ UPDATE [dbo].[order] set [update_date] = '%s',[status_code] = '%s' WHERE [job_num]='%s'""" % (
                    time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())), statuscode, jobnum
                )
                sql = sqlinsertstatus
                objectmssql.ExecNonQuery(sql)
                # returnmsg = "100"
                # 成功的状态被标记，不能重复上传
                if returnmsg == "100":
                    history_list = self.scan_history[_path]
                    if len(history_list) == maxHistory:
                        history_list.pop(0)
                    history_list.append(jobnum)
                return self.build_status_msg(_path, jobnum, returnmsg)
        except Exception as e:
            print(e.__repr__())
            return self.build_status_msg(_path, jobnum, "777")

    def pickingbusinessforprint(self, _path, _data):
        pass

    def buid_out_msg(self, _path, jobnum, stroperation, returnmsg):
        msg = ""
        if returnmsg != '100':
            if returnmsg == "501":
                returnmsg = "镜片已出库"
            if returnmsg == "502":
                returnmsg = "镜架已出库"
            elif returnmsg == "503":
                returnmsg = "无库存"
            elif returnmsg == "504":
                returnmsg = "保留异常"
            elif returnmsg == "505":
                returnmsg = "出库动作异常"
            elif returnmsg == "506":
                returnmsg = "出库单不存在"
            elif returnmsg == "507":
                returnmsg = "生产单不存在"
            elif returnmsg == "508":
                returnmsg = "不需要出库"

            msg = now_time() + "-" + jobnum + "-" + stroperation + "-" + returnmsg
        else:
            returnmsg = "出库成功"

            self.num_recode[_path] += 1

            if stroperation == "镜片":
                # 镜片计量
                stroperation = "片片片"
            else:
                # 镜架计量
                stroperation = "架架架"
            msg = now_time() + "-" + jobnum + "-" + stroperation + "-" + returnmsg
        return msg

    # 构建显示消息，按照一定的格式构建，消息格式参照 deal_data 消息前缀说明
    def build_status_msg(self, _path, jobnum, returnmsg):
        login_value = self.login_recode[_path]
        state_value = self.state_recode[_path]
        if returnmsg == 'SHIP':
            returnmsg = "状态同步失败-订单已完成或取消"
            msg = now_time() + "-" + jobnum + "-" + self.state_recode[_path][0] + "-" + returnmsg
            return "login:%s:%s&state:%s&num:%d&error:%s" % \
                   (login_value[0], login_value[1], state_value[0], self.num_recode[_path], msg)
        elif returnmsg != '100':

            if returnmsg == '502':
                returnmsg = "状态同步失败-工单不存在"
            elif returnmsg == '505':
                returnmsg = "状态同步失败-订单不存在"
            elif returnmsg == '600':
                returnmsg = self.state_recode[_path][0] + "已返工三次及以上"
            elif returnmsg == '999':
                returnmsg = "状态已同步-重复扫描"
            else:
                returnmsg = "状态同步失败-异常"

            msg = now_time() + "-" + jobnum + "-" + self.state_recode[_path][0] + "-" + returnmsg
            return "login:%s:%s&state:%s&num:%d&error:%s" % \
                   (login_value[0], login_value[1], state_value[0], self.num_recode[_path], msg)

        else:
            returnmsg = "状态同步成功"
            self.num_recode[_path] += 1
            msg = now_time() + "-" + jobnum + "-" + self.state_recode[_path][0] + "-" + returnmsg

            return "login:%s:%s&state:%s&num:%d&msg:%s" % \
                   (login_value[0], login_value[1], state_value[0], self.num_recode[_path], msg)

    # 接收结果显示服务器消息的回调，结果显示服务器向代理服务注册其控制的展示面板
    # 注册格式为："r&面板类型&面板唯一路径"
    # 取消注册格式为："ur&面板类型&面板唯一路径"
    # 面板唯一路径："面板ip;面板名"
    def on_result_info_receive(self, data):
        self.deferred[2] = self._init_result_info_deferred()
        try:
            _op, _panel = data.split('&', maxsplit=1)
            if _op == 'r':
                self.output_panel[_panel] = []
                logger.info("register : %s" % _panel)
            elif _op == 'ur':
                for _path in self.output_panel[_panel]:
                    self.login_recode[_path] = None
                    self.state_recode[_path] = None
                    self.num_recode[_path] = 0
                    self.scan_history[_path] = []
                del self.output_panel[_panel]
                logger.info(" un register : %s" % _panel)
            else:
                raise Exception("error op : %s " % _op)
        except Exception as e:
            logger.error("result panel error %s " % e.__repr__())

    def _init_set_deferred(self):
        d = defer.Deferred()
        d.addCallbacks(self.set_result_protocol, self.set_result_failed)
        return d

    def _init_lost_deferred(self):
        d = defer.Deferred()
        d.addBoth(self.on_result_protocol_lost)
        return d

    def _init_result_info_deferred(self):
        d = defer.Deferred()
        d.addBoth(self.on_result_info_receive)
        return d

    def set_result_protocol(self, p):
        self.deferred[0] = self._init_set_deferred()
        self.result_protocol = p
        # print("Set result protocol Success：%s" % self.result_protocol)

    def set_result_failed(self, err):
        self.deferred[0] = self._init_set_deferred()
        # print("Set result protocol failure：%s" % err)

    def on_result_protocol_lost(self, err):
        self.deferred[1] = self._init_lost_deferred()
        self.output_panel = {}
        self.result_protocol = None

    # 向结果服务器发送结果的统一接口
    # 结果格式为："面板类型&面板唯一路径&消息"
    # 消息格式参照 deal_data 消息前缀说明
    def send_result(self, result):
        if self.result_protocol is None:
            logger.error("Result server is not connect")
        else:
            from twisted.internet import reactor
            reactor.callLater(0.01, self.result_protocol.send_result, result)


# 与ResultServer通信Protocol
class ResultClientProtocol(LineReceiver):

    def connectionMade(self):
        self.factory.connect_success(self)

    def lineReceived(self, line):
        self.factory.data_receive(line.decode())

    def send_result(self, result):
        self.sendLine(result.encode())

    def connectionLost(self, reason=main.CONNECTION_LOST):
        self.transport.loseConnection()


# 与ResultServer通信Factory
class ResultClientFactory(ReconnectingClientFactory):
    p = None
    maxDelay = 1800

    def __init__(self, deferred):
        self.deferred = deferred

    def buildProtocol(self, addr):
        self.p = ResultClientProtocol()
        self.p.factory = self
        self.resetDelay()
        # print("Making result protocol ...")
        return self.p

    def data_receive(self, data):
        self.deferred[2].callback(data)

    def connect_success(self, p):
        self.deferred[0].callback(p)
        logger.info("Connected to Result Server and callback to set protocol")

    def clientConnectionFailed(self, connector, reason):
        logger.error("Connect to result server failure, waiting for reconnect...")
        ReconnectingClientFactory.clientConnectionFailed(self, connector, reason)

    def clientConnectionLost(self, connector, unused_reason):
        ReconnectingClientFactory.clientConnectionLost(self, connector, unused_reason)
        logger.error("Lost Connection from result server, waiting for reconnect...")
        self.deferred[1].errback(unused_reason)


def main():
    logger.info("代理服务程序启动")
    factory = ProxyServerFactory()
    from twisted.internet import reactor
    port = reactor.listenTCP(3390, factory)
    logger.info('Proxy Serving transforms on port %d' % port.getHost().port)
    from twisted.internet import reactor
    reactor.run()


import sys
# from PyQt5.QtWidgets import QApplication, QWidget
#
# from proxy_ui import Ui_Form



# class MainUI(QWidget):
#     def __init__(self):
#         super(MainUI, self).__init__()
#         self.Form = Ui_Form()
#         self.Form.setupUi(self)
#         self.Form.start_proxy.clicked.connect(self.start_action)
#         self.Form.stop_proxy.clicked.connect(self.other_action)
#
#     def start_action(self):
#         self.Form.start_proxy.setEnabled(False)
#         self.main()
#
#     def other_action(self):
#         self.Form.stop_proxy.setText("other")
#         pass
#
#     def main(self):
#         factory = ProxyServerFactory()
#         from twisted.internet import reactor
#         port = reactor.listenTCP(3390, factory)
#         print('Proxy Serving transforms on port %d' % port.getHost().port)
#         # from twisted.internet.address import IPv4Address
#         from twisted.internet import reactor
#         reactor.run()

logger = logging.getLogger('Status')
logger.setLevel(logging.DEBUG)
fn = os.path.split(os.path.realpath(__file__))[0] + '/log/' + str(
    time.strftime('%Y-%m-%d', time.localtime(time.time()))) + '.log'
fh = logging.FileHandler(fn, encoding='utf-8')
fh.setLevel(logging.DEBUG)
# ch = logging.StreamHandler()
# ch.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
fh.setFormatter(formatter)
# ch.setFormatter(formatter)
logger.addHandler(fh)
# logger.addHandler(ch)


if __name__ == '__main__':
    main()
    # app = QApplication(sys.argv)
    # # 导入twisted对 PyQt5 的兼容Reactor
    # import qt5reactor
    # qt5reactor.install()
    # w = MainUI()
    # w.setWindowTitle('代理服务实时监控')
    # w.show()
    # sys.exit(app.exec_())
